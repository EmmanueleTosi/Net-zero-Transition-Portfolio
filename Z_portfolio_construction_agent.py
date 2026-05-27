# -*- coding: utf-8 -*-
"""
Portfolio Optimization: Markowitz + Risk Parity with 5-Year Backtest
European Pension Fund — Decarbonization & Energy Transition
Benchmark: MSCI EAFE (EFA)
"""

# Force UTF-8 output
import sys, io
if hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

# ─────────────────────────────────────────────────────────────────────────────
# 0. AUTO-INSTALL DEPENDENCIES
# ─────────────────────────────────────────────────────────────────────────────
import subprocess

def pip_install(package):
    subprocess.check_call([sys.executable, "-m", "pip", "install", package, "--quiet"])

required = {
    "pandas": "pandas", "numpy": "numpy", "openpyxl": "openpyxl",
    "sklearn": "scikit-learn", "cvxpy": "cvxpy",
    "matplotlib": "matplotlib", "yfinance": "yfinance", "scipy": "scipy",
}

print("Checking and installing required libraries...")
for import_name, pkg_name in required.items():
    try:
        __import__(import_name)
    except ImportError:
        print(f"  Installing {pkg_name}...")
        pip_install(pkg_name)
print("All libraries ready.\n")

# ─────────────────────────────────────────────────────────────────────────────
# IMPORTS
# ─────────────────────────────────────────────────────────────────────────────
import warnings
warnings.filterwarnings("ignore")

import pandas as pd
import numpy as np
import cvxpy as cp
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mtick
import matplotlib.dates as mdates
import yfinance as yf
from sklearn.covariance import LedoitWolf
from collections import Counter
from pathlib import Path

# ─────────────────────────────────────────────────────────────────────────────
# STEP 1 — LOAD AND PREPARE DATA
# ─────────────────────────────────────────────────────────────────────────────
print("=" * 68)
print("  STEP 1 — Loading and preparing price data")
print("=" * 68)

df_raw = pd.read_excel("top40_prices.xlsx", engine="openpyxl")
print(f"Raw data loaded: {df_raw.shape[0]} rows x {df_raw.shape[1]} columns")

price_cols = sorted([c for c in df_raw.columns if c.startswith("Price_")])
price_cols = [c for c in price_cols if "Price_2017-01" <= c <= "Price_2026-04"]
print(f"Price columns: {price_cols[0]} -> {price_cols[-1]} ({len(price_cols)} months)")

sectors   = dict(zip(df_raw["ticker"], df_raw["bics_level_1_name"]))
names     = dict(zip(df_raw["ticker"], df_raw["company_name"]))
countries = dict(zip(df_raw["ticker"], df_raw["country_of_incorporation"]))

prices = df_raw.set_index("ticker")[price_cols].T.copy()
prices.index = pd.to_datetime([c.replace("Price_", "") for c in price_cols])
prices.index.name = "date"

threshold = 0.05 * len(prices)
to_drop = prices.isnull().sum()[prices.isnull().sum() > threshold].index.tolist()
if to_drop:
    print(f"Dropping {len(to_drop)} companies with >5% missing: {to_drop}")
    prices = prices.drop(columns=to_drop)

prices  = prices.ffill()
returns = prices.pct_change().dropna(how="all")

tickers       = returns.columns.tolist()
n_companies   = len(tickers)
n_months      = len(returns)
unique_sectors = sorted(set(sectors[t] for t in tickers if t in sectors))

print(f"\nCompanies loaded:   {n_companies}")
print(f"Months (returns):   {n_months}  ({returns.index[0].strftime('%b %Y')} -> {returns.index[-1].strftime('%b %Y')})")
print(f"Sectors ({len(unique_sectors)}): {', '.join(unique_sectors)}")

# ─────────────────────────────────────────────────────────────────────────────
# STEP 2 — DOWNLOAD MSCI EAFE BENCHMARK
# ─────────────────────────────────────────────────────────────────────────────
print("\n" + "=" * 68)
print("  STEP 2 — Downloading MSCI EAFE benchmark (EFA)")
print("=" * 68)

msci_monthly = None
for ticker_bench in ["EFA", "IEFA"]:
    try:
        print(f"  Trying {ticker_bench}...")
        raw = yf.download(ticker_bench, start="2021-01-01", end="2026-05-01",
                          interval="1mo", progress=False, auto_adjust=True)
        if raw.empty:
            raise ValueError("Empty data")
        if isinstance(raw.columns, pd.MultiIndex):
            raw.columns = raw.columns.get_level_values(0)
        close_col = "Close" if "Close" in raw.columns else raw.columns[0]
        msci_prices  = raw[close_col].dropna()
        msci_monthly = msci_prices.pct_change().dropna()
        msci_monthly.index = msci_monthly.index.to_period("M").to_timestamp()
        print(f"  {ticker_bench} downloaded: {len(msci_monthly)} monthly returns "
              f"({msci_monthly.index[0].strftime('%b %Y')} -> {msci_monthly.index[-1].strftime('%b %Y')})")
        BENCH_TICKER = ticker_bench
        break
    except Exception as e:
        print(f"  {ticker_bench} failed: {e}")

if msci_monthly is None:
    print("  WARNING: benchmark unavailable — using zero returns as fallback.")
    msci_monthly = pd.Series(0.0, index=pd.date_range("2021-02-01", periods=63, freq="MS"))
    BENCH_TICKER = "N/A"

# ─────────────────────────────────────────────────────────────────────────────
# STEP 3 — ROLLING WINDOW STRUCTURE
# ─────────────────────────────────────────────────────────────────────────────
print("\n" + "=" * 68)
print("  STEP 3 — Defining rolling window structure")
print("=" * 68)

backtest_windows = [
    dict(iter=1, tr_s=0,  tr_e=48,  te_s=48,  te_e=60,
         label="Iter 1: Train Jan2017-Dec2020 | Test Jan2021-Dec2021"),
    dict(iter=2, tr_s=12, tr_e=60,  te_s=60,  te_e=72,
         label="Iter 2: Train Jan2018-Dec2021 | Test Jan2022-Dec2022"),
    dict(iter=3, tr_s=24, tr_e=72,  te_s=72,  te_e=84,
         label="Iter 3: Train Jan2019-Dec2022 | Test Jan2023-Dec2023"),
    dict(iter=4, tr_s=36, tr_e=84,  te_s=84,  te_e=96,
         label="Iter 4: Train Jan2020-Dec2023 | Test Jan2024-Dec2024"),
    dict(iter=5, tr_s=51, tr_e=99,  te_s=99,  te_e=111,
         label="Iter 5: Train May2021-Apr2025 | Test May2025-Apr2026"),
]
current_window = dict(tr_s=63, tr_e=111, label="Current: Train May2022-Apr2026")

for w in backtest_windows:
    print(f"  {w['label']}")
print(f"  {current_window['label']}")

# ─────────────────────────────────────────────────────────────────────────────
# STEP 4A — MARKOWITZ OPTIMIZATION
# ─────────────────────────────────────────────────────────────────────────────

def optimize_markowitz(train_returns, sectors_dict, rf=0.025, lambda_risk=3.0):
    """Maximize Sharpe via mean-variance (convex QP, Ledoit-Wolf shrinkage)."""
    tick = train_returns.columns.tolist()
    n    = len(tick)

    # 4A.1 Expected returns
    mean_monthly  = train_returns.mean()
    expected_annual = (1 + mean_monthly) ** 12 - 1

    # 4A.2 Covariance — Ledoit-Wolf
    lw = LedoitWolf()
    lw.fit(train_returns.values)
    cov_annual = lw.covariance_ * 12
    cov_annual = (cov_annual + cov_annual.T) / 2
    eig = np.linalg.eigvalsh(cov_annual).min()
    if eig < 0:
        cov_annual += (-eig + 1e-8) * np.eye(n)

    # 4A.3 Convex optimization
    w  = cp.Variable(n, nonneg=True)
    mu = expected_annual.values

    sec_list   = [sectors_dict.get(t, "Unknown") for t in tick]
    sec_counts = Counter(sec_list)
    top5       = [s for s, _ in sec_counts.most_common(5)]
    sec_cons   = []
    for sec in top5:
        idx = [i for i, t in enumerate(tick) if sectors_dict.get(t) == sec]
        if idx:
            sec_cons.append(cp.sum(w[idx]) >= 0.005)

    constraints = [cp.sum(w) == 1, w >= 0.005, w <= 0.10] + sec_cons
    objective   = cp.Maximize(mu @ w - lambda_risk * cp.quad_form(w, cov_annual))
    problem     = cp.Problem(objective, constraints)

    solved = False
    for solver in [cp.CLARABEL, cp.ECOS, cp.SCS]:
        try:
            problem.solve(solver=solver, verbose=False)
            if w.value is not None and problem.status in ["optimal", "optimal_inaccurate"]:
                solved = True
                break
        except Exception:
            continue

    if not solved:
        return None

    wa = np.clip(np.array(w.value, dtype=float), 0, None)
    wa[wa < 0.005] = 0
    if wa.sum() < 1e-8:
        return None
    wa /= wa.sum()
    # Iterative cap: renormalizing after capping can push weights back above 10%,
    # so repeat until every weight is strictly within the 10% limit.
    for _ in range(100):
        wa = np.minimum(wa, 0.10)
        wa /= wa.sum()
        if wa.max() <= 0.1001:
            break

    weights    = pd.Series(wa, index=tick)
    port_ret   = float(mu @ wa)
    port_vol   = float(np.sqrt(wa @ cov_annual @ wa))
    sharpe     = (port_ret - rf) / port_vol if port_vol > 0 else 0.0
    n_sel      = int((wa > 1e-4).sum())
    active_sec = set(sectors_dict.get(t) for t, ww in zip(tick, wa) if ww > 1e-4)

    return dict(weights=weights, expected_return=port_ret, volatility=port_vol,
                sharpe_ratio=sharpe, n_companies=n_sel, n_sectors=len(active_sec))

# ─────────────────────────────────────────────────────────────────────────────
# STEP 4B — RISK PARITY OPTIMIZATION
# ─────────────────────────────────────────────────────────────────────────────

def optimize_risk_parity(train_returns, rf=0.025):
    """Log-barrier Risk Parity on all companies (convex DCP).

    min  0.5 * w'Sigma*w  -  (1/n)*sum(log(w))
    At optimum: w_i*(Sigma w)_i = const  =>  equal risk contribution.
    """
    tick = train_returns.columns.tolist()
    n    = len(tick)

    lw = LedoitWolf()
    lw.fit(train_returns.values)
    cov_annual = lw.covariance_ * 12
    cov_annual = (cov_annual + cov_annual.T) / 2
    eig = np.linalg.eigvalsh(cov_annual).min()
    if eig < 0:
        cov_annual += (-eig + 1e-8) * np.eye(n)

    w         = cp.Variable(n)
    objective = cp.Minimize(0.5 * cp.quad_form(w, cov_annual) - (1.0/n) * cp.sum(cp.log(w)))
    problem   = cp.Problem(objective, [w >= 1e-6])

    solved = False
    for solver in [cp.CLARABEL, cp.ECOS, cp.SCS]:
        try:
            problem.solve(solver=solver, verbose=False)
            if w.value is not None and problem.status in ["optimal", "optimal_inaccurate"]:
                solved = True
                break
        except Exception:
            continue

    if not solved:
        return None

    wa = np.clip(np.array(w.value, dtype=float), 0, None)
    wa[wa < 0.001] = 0
    if wa.sum() < 1e-8:
        return None
    wa /= wa.sum()

    weights         = pd.Series(wa, index=tick)
    mean_monthly    = train_returns.mean()
    expected_annual = (1 + mean_monthly) ** 12 - 1
    port_ret  = float(expected_annual.values @ wa)
    port_vol  = float(np.sqrt(wa @ cov_annual @ wa))
    sharpe    = (port_ret - rf) / port_vol if port_vol > 0 else 0.0
    n_sel     = int((wa > 1e-4).sum())
    active_sec = set(sectors.get(t) for t, ww in zip(tick, wa) if ww > 1e-4)

    return dict(weights=weights, expected_return=port_ret, volatility=port_vol,
                sharpe_ratio=sharpe, n_companies=n_sel, n_sectors=len(active_sec))

# ─────────────────────────────────────────────────────────────────────────────
# STEP 5 — RUN BACKTEST
# ─────────────────────────────────────────────────────────────────────────────
print("\n" + "=" * 68)
print("  STEP 5 — Running 5-iteration backtest (Markowitz & Risk Parity)")
print("=" * 68)

def compute_metrics(monthly_ret, rf=0.025):
    n   = len(monthly_ret)
    cum = (1 + monthly_ret).prod() - 1
    ann_ret = (1 + cum) ** (12 / n) - 1
    ann_vol = monthly_ret.std() * np.sqrt(12)
    sharpe  = (ann_ret - rf) / ann_vol if ann_vol > 0 else 0.0
    curve   = (1 + monthly_ret).cumprod()
    max_dd  = ((curve - curve.cummax()) / curve.cummax()).min()
    return dict(cumulative=cum, ann_return=ann_ret, ann_vol=ann_vol,
                sharpe=sharpe, max_drawdown=max_dd)

backtest_records = []
mkw_monthly_all  = []
rp_monthly_all   = []
msci_mkw_all     = []
msci_rp_all      = []

returns_ms = returns.copy()
returns_ms.index = returns_ms.index.to_period("M").to_timestamp()

for w_def in backtest_windows:
    it = w_def["iter"]
    print(f"\n  Iteration {it}: {w_def['label']}")

    train_ret  = returns_ms.iloc[w_def["tr_s"]: w_def["tr_e"]]
    test_ret   = returns_ms.iloc[w_def["te_s"]: w_def["te_e"]]
    test_dates = test_ret.index
    msci_test  = msci_monthly.reindex(test_dates).fillna(0)

    tr_s = train_ret.index[0].strftime("%Y-%m")
    tr_e = train_ret.index[-1].strftime("%Y-%m")
    te_s = test_ret.index[0].strftime("%Y-%m")
    te_e = test_ret.index[-1].strftime("%Y-%m")

    # Markowitz
    print(f"    Optimizing Markowitz...", end=" ")
    mkw = optimize_markowitz(train_ret, sectors)
    if mkw is None:
        print("FAILED")
    else:
        print(f"OK  ({mkw['n_companies']} companies, {mkw['n_sectors']} sectors)")
        common  = [t for t in mkw["weights"].index if t in test_ret.columns]
        test_m  = (test_ret[common] * mkw["weights"][common].values).sum(axis=1)
        test_m.index = test_dates
        met_m   = compute_metrics(test_m)
        mkw_monthly_all.append(test_m)
        msci_mkw_all.append(msci_test)
        backtest_records.append(dict(
            Method="Markowitz", Iteration=it,
            Train_Start=tr_s, Train_End=tr_e, Test_Start=te_s, Test_End=te_e,
            Ann_Return=f"{met_m['ann_return']:.1%}", Ann_Vol=f"{met_m['ann_vol']:.1%}",
            Sharpe=f"{met_m['sharpe']:.2f}", Max_Drawdown=f"{met_m['max_drawdown']:.1%}",
            N_Companies=mkw["n_companies"], N_Sectors=mkw["n_sectors"],
        ))

    # Risk Parity
    print(f"    Optimizing Risk Parity...", end=" ")
    rp = optimize_risk_parity(train_ret)
    if rp is None:
        print("FAILED")
    else:
        print(f"OK  ({rp['n_companies']} companies, {rp['n_sectors']} sectors)")
        common  = [t for t in rp["weights"].index if t in test_ret.columns]
        test_r  = (test_ret[common] * rp["weights"][common].values).sum(axis=1)
        test_r.index = test_dates
        met_r   = compute_metrics(test_r)
        rp_monthly_all.append(test_r)
        msci_rp_all.append(msci_test)
        backtest_records.append(dict(
            Method="Risk Parity", Iteration=it,
            Train_Start=tr_s, Train_End=tr_e, Test_Start=te_s, Test_End=te_e,
            Ann_Return=f"{met_r['ann_return']:.1%}", Ann_Vol=f"{met_r['ann_vol']:.1%}",
            Sharpe=f"{met_r['sharpe']:.2f}", Max_Drawdown=f"{met_r['max_drawdown']:.1%}",
            N_Companies=rp["n_companies"], N_Sectors=rp["n_sectors"],
        ))

mkw_full  = pd.concat(mkw_monthly_all) if mkw_monthly_all else pd.Series(dtype=float)
rp_full   = pd.concat(rp_monthly_all)  if rp_monthly_all  else pd.Series(dtype=float)
msci_mkw  = pd.concat(msci_mkw_all)    if msci_mkw_all    else pd.Series(dtype=float)
msci_rp   = pd.concat(msci_rp_all)     if msci_rp_all     else pd.Series(dtype=float)

# ─────────────────────────────────────────────────────────────────────────────
# STEP 6 — CURRENT PORTFOLIO
# ─────────────────────────────────────────────────────────────────────────────
print("\n" + "=" * 68)
print("  STEP 6 — Building current portfolio (May 2022 - Apr 2026)")
print("=" * 68)

cur_train = returns_ms.iloc[current_window["tr_s"]: current_window["tr_e"]]

print("  Optimizing Markowitz (current)...", end=" ")
cur_mkw = optimize_markowitz(cur_train, sectors)
print(f"OK  ({cur_mkw['n_companies']} companies, vol={cur_mkw['volatility']:.1%})" if cur_mkw else "FAILED")

print("  Optimizing Risk Parity (current)...", end=" ")
cur_rp = optimize_risk_parity(cur_train)
print(f"OK  ({cur_rp['n_companies']} companies, vol={cur_rp['volatility']:.1%})" if cur_rp else "FAILED")

# ─────────────────────────────────────────────────────────────────────────────
# STEP 7 — BETA vs MSCI EAFE
# ─────────────────────────────────────────────────────────────────────────────
print("\n" + "=" * 68)
print("  STEP 7 — Computing beta vs MSCI EAFE")
print("=" * 68)

def compute_beta(port_s, bench_s):
    df = pd.concat([port_s, bench_s], axis=1).dropna()
    if len(df) < 2:
        return float("nan")
    p, b = df.iloc[:, 0].values, df.iloc[:, 1].values
    return np.cov(p, b)[0, 1] / np.var(b) if np.var(b) > 0 else float("nan")

beta_mkw = compute_beta(mkw_full, msci_mkw)
beta_rp  = compute_beta(rp_full,  msci_rp)
print(f"  Markowitz beta vs {BENCH_TICKER}:   {beta_mkw:.2f}")
print(f"  Risk Parity beta vs {BENCH_TICKER}: {beta_rp:.2f}")

# ─────────────────────────────────────────────────────────────────────────────
# STEP 8 — OUTPUTS
# ─────────────────────────────────────────────────────────────────────────────
print("\n" + "=" * 68)
print("  STEP 8 — Producing outputs")
print("=" * 68)

_EMPTY = dict(cumulative=0.0, ann_return=0.0, ann_vol=0.0, sharpe=0.0, max_drawdown=0.0)
met_mkw  = compute_metrics(mkw_full) if len(mkw_full) > 0 else _EMPTY
met_rp   = compute_metrics(rp_full)  if len(rp_full)  > 0 else _EMPTY
met_mscim = compute_metrics(msci_mkw) if len(msci_mkw) > 0 else _EMPTY
met_mscir = compute_metrics(msci_rp)  if len(msci_rp)  > 0 else _EMPTY
# Use MKW-aligned MSCI for the shared summary table
met_msci = met_mscim

SEP  = "=" * 68
sep2 = "-" * 68

# OUTPUT 3 — Summary table
print(f"\n{SEP}")
print("  BACKTEST PERFORMANCE SUMMARY (Feb 2021 - Apr 2026)")
print(SEP)
print(f"{'':24s} {'Markowitz':>12} {'Risk Parity':>13} {'MSCI EAFE':>12}")
print(sep2)
print(f"  {'Total return 5yr:':<22} {met_mkw['cumulative']:>+11.1%} {met_rp['cumulative']:>+12.1%} {met_msci['cumulative']:>+11.1%}")
print(f"  {'Annualized return:':<22} {met_mkw['ann_return']:>11.1%} {met_rp['ann_return']:>12.1%} {met_msci['ann_return']:>11.1%}")
print(f"  {'Annualized volatility:':<22} {met_mkw['ann_vol']:>11.1%} {met_rp['ann_vol']:>12.1%} {met_msci['ann_vol']:>11.1%}")
print(f"  {'Sharpe Ratio:':<22} {met_mkw['sharpe']:>11.2f} {met_rp['sharpe']:>12.2f} {met_msci['sharpe']:>11.2f}")
print(f"  {'Beta vs MSCI EAFE:':<22} {beta_mkw:>11.2f} {beta_rp:>12.2f} {'1.00':>11}")
print(f"  {'Max drawdown:':<22} {met_mkw['max_drawdown']:>+11.1%} {met_rp['max_drawdown']:>+12.1%} {met_msci['max_drawdown']:>+11.1%}")
print(SEP)

# OUTPUT 4 — Current portfolio detail
def print_portfolio(title, res, check_constraints=False):
    if res is None:
        print(f"\n{title}: optimization failed.")
        return
    active = res["weights"][res["weights"] > 1e-4].sort_values(ascending=False)
    print(f"\n{SEP}")
    print(f"  {title}")
    print(SEP)
    print(f"  {'Rank':<5} {'Ticker':<8} {'Company Name':<30} {'Sector':<22} {'Country':<8} {'Weight':>7}")
    print(sep2)
    for rank, (t, ww) in enumerate(active.items(), 1):
        print(f"  {rank:<5} {t:<8} {names.get(t,t)[:28]:<30} {sectors.get(t,'?')[:20]:<22} {countries.get(t,'?')[:6]:<8} {ww:>6.1%}")
    print(sep2)
    print(f"  Expected annual return:  {res['expected_return']:>6.1%}")
    print(f"  Annual volatility:       {res['volatility']:>6.1%}")
    print(f"  Sharpe Ratio:            {res['sharpe_ratio']:>6.2f}")
    print(f"  Number of companies:     {res['n_companies']}")
    print(f"  Number of sectors:       {res['n_sectors']}")
    print(f"  Max single weight:       {active.max():>6.1%}")

    sec_alloc = {}
    for t, ww in active.items():
        s = sectors.get(t, "Other")
        sec_alloc[s] = sec_alloc.get(s, 0) + ww
    print(f"\n  Sector allocation:")
    for s, ww in sorted(sec_alloc.items(), key=lambda x: -x[1]):
        print(f"    {s:<32} {ww:>6.1%}")

    cty_alloc = {}
    for t, ww in active.items():
        c = countries.get(t, "Other")
        cty_alloc[c] = cty_alloc.get(c, 0) + ww
    print(f"\n  Country allocation:")
    for c, ww in sorted(cty_alloc.items(), key=lambda x: -x[1]):
        print(f"    {c:<32} {ww:>6.1%}")

    if check_constraints:
        ok = lambda cond: "OK" if cond else "!!"
        print(f"\n  CONSTRAINT CHECK (informational):")
        print(f"    Max weight:         {active.max():>6.1%}  [limit 10%  -> {ok(active.max() <= 0.10)}]")
        print(f"    Number companies:   {res['n_companies']:>4}     [no limit   -> {ok(True)}]")
        print(f"    Number sectors:     {res['n_sectors']:>4}     [minimum 5  -> {ok(res['n_sectors'] >= 5)}]")
        print(f"    Portfolio vol:      {res['volatility']:>6.1%}  [limit 12%  -> {ok(res['volatility'] <= 0.12)}]")
    print(SEP)

print_portfolio("CURRENT PORTFOLIO — MARKOWITZ (trained May 2022 - Apr 2026)", cur_mkw)
print_portfolio("CURRENT PORTFOLIO — RISK PARITY (trained May 2022 - Apr 2026)", cur_rp, check_constraints=True)

# OUTPUT 5 — Top 5 holdings
print(f"\n{SEP}")
print("  TOP 5 HOLDINGS")
print(SEP)
print(f"  {'Rank':<5} {'Method':<14} {'Ticker':<8} {'Company Name':<28} {'Sector':<22} {'Weight':>7}")
print(sep2)
rank = 1
for res_obj, method_lbl in [(cur_mkw, "Markowitz"), (cur_rp, "Risk Parity")]:
    if res_obj:
        for t, ww in res_obj["weights"].sort_values(ascending=False).head(5).items():
            print(f"  {rank:<5} {method_lbl:<14} {t:<8} {names.get(t,t)[:26]:<28} {sectors.get(t,'?')[:20]:<22} {ww:>6.1%}")
            rank += 1
    if method_lbl == "Markowitz":
        print(sep2)
print(SEP)

# OUTPUT 1 — Excel
print("\n  Saving portfolio_results.xlsx...")
mkw_w = cur_mkw["weights"] if cur_mkw else pd.Series(0.0, index=tickers)
rp_w  = cur_rp["weights"]  if cur_rp  else pd.Series(0.0, index=tickers)
df_out = df_raw.copy()
df_out.insert(0, "Weight_RiskParity", df_out["ticker"].map(rp_w).fillna(0.0))
df_out.insert(0, "Weight_Markowitz",  df_out["ticker"].map(mkw_w).fillna(0.0))
with pd.ExcelWriter("Z_portfolio_results.xlsx", engine="openpyxl") as writer:
    df_out.to_excel(writer, index=False, sheet_name="Portfolio")
    ws = writer.sheets["Portfolio"]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
        for cell in row:
            cell.number_format = "0.0%"
print("  portfolio_results.xlsx saved.")

# OUTPUT 6 — CSV
print("  Saving backtest_results.csv...")
pd.DataFrame(backtest_records).to_csv("Z_backtest_results.csv", index=False)
print("  backtest_results.csv saved.")

# OUTPUT — Monthly returns (for Z_comparison_agent.py)
print("  Saving Z_monthly_returns.xlsx...")
pd.DataFrame({
    "Markowitz": mkw_full,
    "RiskParity": rp_full,
    "MSCI_EAFE":  msci_mkw,
}).sort_index().rename_axis("Date").to_excel("Z_monthly_returns.xlsx", engine="openpyxl")
print("  Z_monthly_returns.xlsx saved.")

# OUTPUT 2 — Charts
print("  Generating performance charts...")

def build_chart(port_ret, bench_ret, port_label, port_color, filename, title,
                met_port, met_bench, beta_val):
    df = pd.concat([port_ret, bench_ret], axis=1).dropna()
    df.columns = ["port", "bench"]
    cum_p = (1 + df["port"]).cumprod()  * 100
    cum_b = (1 + df["bench"]).cumprod() * 100

    try:
        plt.style.use("seaborn-v0_8-whitegrid")
    except Exception:
        try:
            plt.style.use("seaborn-whitegrid")
        except Exception:
            pass

    fig, ax = plt.subplots(figsize=(12, 7))
    ax.plot(cum_p.index, cum_p.values, color=port_color, linewidth=2.5, label=port_label, zorder=3)
    ax.plot(cum_b.index, cum_b.values, color="#888888", linewidth=1.8, linestyle="--",
            label=f"MSCI EAFE ({BENCH_TICKER})", zorder=2)
    ax.axhline(100, color="black", linewidth=0.6, linestyle=":", alpha=0.5)
    ax.set_title(title, fontsize=14, fontweight="bold", pad=15)
    ax.set_xlabel("Date", fontsize=11)
    ax.set_ylabel("Cumulative Value (Base = 100)", fontsize=11)
    ax.legend(fontsize=11, loc="upper left")
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%b %Y"))
    fig.autofmt_xdate(rotation=30)
    ax.yaxis.set_major_formatter(mtick.FormatStrFormatter("%.0f"))

    textstr = (
        f"Portfolio total return: {met_port['cumulative']:+.1%}\n"
        f"MSCI EAFE total return: {met_bench['cumulative']:+.1%}\n"
        f"Portfolio volatility: {met_port['ann_vol']:.1%}\n"
        f"Sharpe Ratio: {met_port['sharpe']:.2f}\n"
        f"Beta: {beta_val:.2f}"
    )
    ax.text(0.98, 0.05, textstr, transform=ax.transAxes, fontsize=9,
            verticalalignment="bottom", horizontalalignment="right",
            bbox=dict(boxstyle="round,pad=0.5", facecolor="white", alpha=0.85, edgecolor="#cccccc"))
    plt.tight_layout()
    plt.savefig(filename, dpi=150, bbox_inches="tight")
    plt.close()
    print(f"  {filename} saved.")

build_chart(mkw_full, msci_mkw, "Markowitz Portfolio", "#2ecc71",
            "Z_chart_markowitz_vs_msci.png",
            "Markowitz Portfolio vs MSCI EAFE — 5 Year Backtest (Feb 2021 - Apr 2026)",
            met_mkw, met_mscim, beta_mkw)

build_chart(rp_full, msci_rp, "Risk Parity Portfolio", "#3498db",
            "Z_chart_riskparity_vs_msci.png",
            "Risk Parity Portfolio vs MSCI EAFE — 5 Year Backtest (Feb 2021 - Apr 2026)",
            met_rp, met_mscir, beta_rp)

# Markowitz-only chart (no benchmark)
def build_markowitz_only_chart(port_ret, port_label, port_color, filename, title, met_port):
    cum_p = (1 + port_ret).cumprod() * 100

    try:
        plt.style.use("seaborn-v0_8-whitegrid")
    except Exception:
        try:
            plt.style.use("seaborn-whitegrid")
        except Exception:
            pass

    fig, ax = plt.subplots(figsize=(12, 7))
    ax.plot(cum_p.index, cum_p.values, color=port_color, linewidth=2.5, label=port_label, zorder=3)
    ax.axhline(100, color="black", linewidth=0.6, linestyle=":", alpha=0.5)
    ax.set_title(title, fontsize=14, fontweight="bold", pad=15)
    ax.set_xlabel("Date", fontsize=11)
    ax.set_ylabel("Cumulative Value (Base = 100)", fontsize=11)
    ax.legend(fontsize=11, loc="upper left")
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%b %Y"))
    fig.autofmt_xdate(rotation=30)
    ax.yaxis.set_major_formatter(mtick.FormatStrFormatter("%.0f"))

    textstr = (
        f"Total return: {met_port['cumulative']:+.1%}\n"
        f"Annualized return: {met_port['ann_return']:.1%}\n"
        f"Annualized volatility: {met_port['ann_vol']:.1%}\n"
        f"Sharpe Ratio: {met_port['sharpe']:.2f}\n"
        f"Max drawdown: {met_port['max_drawdown']:+.1%}"
    )
    ax.text(0.98, 0.05, textstr, transform=ax.transAxes, fontsize=9,
            verticalalignment="bottom", horizontalalignment="right",
            bbox=dict(boxstyle="round,pad=0.5", facecolor="white", alpha=0.85, edgecolor="#cccccc"))
    plt.tight_layout()
    plt.savefig(filename, dpi=150, bbox_inches="tight")
    plt.close()
    print(f"  {filename} saved.")

build_markowitz_only_chart(mkw_full, "Markowitz Portfolio", "#2ecc71",
                           "Z_chart_markowitz_only.png",
                           "Markowitz Portfolio — 5 Year Backtest (Feb 2021 - Apr 2026)",
                           met_mkw)

# ─────────────────────────────────────────────────────────────────────────────
# STEP 9 — METHODOLOGY SUMMARY
# ─────────────────────────────────────────────────────────────────────────────
print(f"\n{SEP}")
print("  METHODOLOGY")
print(SEP)
print(f"""Method 1 - Markowitz Mean-Variance Optimization:
  Covariance estimation: Ledoit-Wolf shrinkage (Ledoit & Wolf, 2004)
  Expected returns: geometric annualization of 48-month mean monthly returns
  Objective: maximize risk-adjusted return (mean-variance formulation)
  Constraints: max weight 10%, min weight 0.5%,
               at least 5 sectors each receiving minimum 0.5% allocation
  Solver: CLARABEL (convex QP, no binary variables)

Method 2 - Risk Parity:
  Covariance estimation: Ledoit-Wolf shrinkage (same as Markowitz)
  Objective: equalize risk contribution of each asset (log-barrier formulation)
  Constraints: weights sum to 1, minimum 0.1% per asset (numerical stability)
  No explicit constraints on weight, sectors, or number of companies
  Solver: CLARABEL

Backtest:
  Rolling window, annual rebalancing — 5 iterations
  Training window: 48 months | Test window: 12 months
  No look-ahead bias: each portfolio uses only data available at time t
  Benchmark: MSCI EAFE ETF (ticker: {BENCH_TICKER})
  Risk-free rate: 2.5% (German Bund approximation)

Limitation:
  ESG universe fixed at current screening (no point-in-time historical ESG data)
  Expected returns estimated from historical means — not a guarantee of future performance""")
print(SEP)

# ─────────────────────────────────────────────────────────────────────────────
# OUTPUT 7 — COMPARISON EXCEL FILE
# ─────────────────────────────────────────────────────────────────────────────
print("\n  Saving comparison_report.xlsx...")

from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Colour palette ────────────────────────────────────────────────────────
C_MKW   = "C6EFCE"   # light green  — Markowitz
C_RP    = "BDD7EE"   # light blue   — Risk Parity
C_MSCI  = "E2EFDA"   # light grey-green — MSCI EAFE
C_HEAD  = "1F4E79"   # dark navy    — header background
C_TITLE = "2E75B6"   # mid blue     — section titles
C_ALT   = "F2F2F2"   # light grey   — alternating rows

def hdr_font(bold=True, color="FFFFFF", sz=11):
    return Font(bold=bold, color=color, size=sz)

def cell_font(bold=False, color="000000", sz=10):
    return Font(bold=bold, color=color, size=sz)

def fill(hex_col):
    return PatternFill("solid", fgColor=hex_col)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def pct(val, decimals=1):
    """Format float as percentage string."""
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return "N/A"
    return f"{val:+.{decimals}%}" if val < 0 or val >= 0 else f"{val:.{decimals}%}"

def write_header_row(ws, row, values, bg=C_HEAD, font_color="FFFFFF", sz=11):
    for col, v in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=v)
        c.font      = Font(bold=True, color=font_color, size=sz)
        c.fill      = fill(bg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = thin_border()

def write_data_row(ws, row, values, bg=None, bold=False, num_fmts=None, aligns=None):
    for col, v in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=v)
        c.font      = Font(bold=bold, size=10)
        c.alignment = Alignment(horizontal=aligns[col-1] if aligns else "center",
                                vertical="center")
        c.border    = thin_border()
        if bg:
            c.fill = fill(bg)
        if num_fmts and num_fmts[col-1]:
            c.number_format = num_fmts[col-1]

# ══════════════════════════════════════════════════════════════════════════
# SHEET 1 — PERFORMANCE SUMMARY
# ══════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Performance Summary"
ws1.sheet_view.showGridLines = False
ws1.column_dimensions["A"].width = 28
ws1.column_dimensions["B"].width = 16
ws1.column_dimensions["C"].width = 16
ws1.column_dimensions["D"].width = 16

# Title
ws1.merge_cells("A1:D1")
t = ws1["A1"]
t.value     = "Portfolio Comparison vs MSCI EAFE — 5-Year Backtest (Feb 2021 – Apr 2026)"
t.font      = Font(bold=True, size=13, color="FFFFFF")
t.fill      = fill(C_HEAD)
t.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 30

write_header_row(ws1, 2, ["Metric", "Markowitz", "Risk Parity", f"MSCI EAFE ({BENCH_TICKER})"],
                 bg=C_TITLE)
ws1.row_dimensions[2].height = 22

rows_data = [
    ("Total Return (5yr)",       met_mkw["cumulative"],    met_rp["cumulative"],    met_msci["cumulative"]),
    ("Annualized Return",        met_mkw["ann_return"],    met_rp["ann_return"],    met_msci["ann_return"]),
    ("Annualized Volatility",    met_mkw["ann_vol"],       met_rp["ann_vol"],       met_msci["ann_vol"]),
    ("Sharpe Ratio",             met_mkw["sharpe"],        met_rp["sharpe"],        met_msci["sharpe"]),
    (f"Beta vs MSCI EAFE",       beta_mkw,                 beta_rp,                 1.00),
    ("Max Drawdown",             met_mkw["max_drawdown"],  met_rp["max_drawdown"],  met_msci["max_drawdown"]),
    ("Ann. Return / Volatility", met_mkw["ann_return"]/met_mkw["ann_vol"] if met_mkw["ann_vol"] else 0,
                                  met_rp["ann_return"]/met_rp["ann_vol"]   if met_rp["ann_vol"]  else 0,
                                  met_msci["ann_return"]/met_msci["ann_vol"] if met_msci["ann_vol"] else 0),
]

pct_rows   = {0, 1, 2, 5}    # indices that are percentages
ratio_rows = {3, 4, 6}        # indices that are plain numbers

for i, (label, v_mkw, v_rp, v_msci) in enumerate(rows_data):
    r    = i + 3
    bg_r = C_ALT if i % 2 == 0 else "FFFFFF"

    def fmt(v, idx=i):
        if idx in pct_rows:
            return f"{v:+.1%}" if idx in {0, 5} else f"{v:.1%}"
        return f"{v:.2f}"

    write_data_row(ws1, r,
                   [label, fmt(v_mkw), fmt(v_rp), fmt(v_msci)],
                   bg=bg_r,
                   aligns=["left", "center", "center", "center"])
    # Colour the portfolio cells
    ws1.cell(r, 2).fill = fill(C_MKW)
    ws1.cell(r, 3).fill = fill(C_RP)
    ws1.cell(r, 4).fill = fill(C_MSCI)

# Divider row
r_div = len(rows_data) + 3
ws1.merge_cells(f"A{r_div}:D{r_div}")
ws1[f"A{r_div}"].fill = fill(C_HEAD)
ws1.row_dimensions[r_div].height = 6

# Current portfolio stats
r_cur = r_div + 1
write_header_row(ws1, r_cur,
                 ["Current Portfolio Stats", "Markowitz", "Risk Parity", ""],
                 bg=C_TITLE)
ws1.row_dimensions[r_cur].height = 22

cur_rows = [
    ("Number of Companies",
     cur_mkw["n_companies"] if cur_mkw else "-",
     cur_rp["n_companies"]  if cur_rp  else "-", ""),
    ("Number of Sectors",
     cur_mkw["n_sectors"]   if cur_mkw else "-",
     cur_rp["n_sectors"]    if cur_rp  else "-", ""),
    ("Expected Annual Return",
     f"{cur_mkw['expected_return']:.1%}" if cur_mkw else "-",
     f"{cur_rp['expected_return']:.1%}"  if cur_rp  else "-", ""),
    ("Portfolio Volatility",
     f"{cur_mkw['volatility']:.1%}"      if cur_mkw else "-",
     f"{cur_rp['volatility']:.1%}"       if cur_rp  else "-", ""),
    ("Max Single Weight",
     f"{cur_mkw['weights'].max():.1%}"   if cur_mkw else "-",
     f"{cur_rp['weights'].max():.1%}"    if cur_rp  else "-", ""),
]
for i, row_vals in enumerate(cur_rows):
    r   = r_cur + 1 + i
    bg_r = C_ALT if i % 2 == 0 else "FFFFFF"
    write_data_row(ws1, r, list(row_vals), bg=bg_r,
                   aligns=["left", "center", "center", "center"])
    ws1.cell(r, 2).fill = fill(C_MKW)
    ws1.cell(r, 3).fill = fill(C_RP)

# Legend
r_leg = r_cur + len(cur_rows) + 2
ws1.merge_cells(f"A{r_leg}:D{r_leg}")
leg = ws1[f"A{r_leg}"]
leg.value     = "Green = Markowitz   |   Blue = Risk Parity   |   Light grey-green = MSCI EAFE"
leg.font      = Font(italic=True, size=9, color="666666")
leg.alignment = Alignment(horizontal="center")

# ══════════════════════════════════════════════════════════════════════════
# SHEET 2 — BACKTEST BY ITERATION
# ══════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Backtest by Iteration")
ws2.sheet_view.showGridLines = False
for col, w in zip("ABCDEFGHIJ", [12,12,12,12,12,12,12,12,14,12]):
    ws2.column_dimensions[get_column_letter(ord(col)-64)].width = w

ws2.merge_cells("A1:J1")
t2 = ws2["A1"]
t2.value     = "Backtest Results by Iteration"
t2.font      = Font(bold=True, size=13, color="FFFFFF")
t2.fill      = fill(C_HEAD)
t2.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 30

hdrs2 = ["Method", "Iteration", "Train Start", "Train End",
          "Test Start", "Test End", "Ann. Return", "Ann. Vol",
          "Sharpe", "Max Drawdown"]
write_header_row(ws2, 2, hdrs2, bg=C_TITLE)

for i, rec in enumerate(backtest_records):
    r   = i + 3
    bg_r = C_MKW if rec["Method"] == "Markowitz" else C_RP
    vals = [rec["Method"], rec["Iteration"],
            rec["Train_Start"], rec["Train_End"],
            rec["Test_Start"],  rec["Test_End"],
            rec["Ann_Return"],  rec["Ann_Vol"],
            rec["Sharpe"],      rec["Max_Drawdown"]]
    write_data_row(ws2, r, vals, bg=bg_r,
                   aligns=["left","center","center","center",
                            "center","center","center","center","center","center"])

# ══════════════════════════════════════════════════════════════════════════
# SHEET 3 — MONTHLY RETURNS & CUMULATIVE
# ══════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Monthly Returns")
ws3.sheet_view.showGridLines = False
for col, w in zip(range(1,8), [14,12,12,12,14,14,14]):
    ws3.column_dimensions[get_column_letter(col)].width = w

ws3.merge_cells("A1:G1")
t3 = ws3["A1"]
t3.value     = "Monthly & Cumulative Returns — Backtest Period"
t3.font      = Font(bold=True, size=13, color="FFFFFF")
t3.fill      = fill(C_HEAD)
t3.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 30

write_header_row(ws3, 2,
    ["Date", "Markowitz\nMonthly", "Risk Parity\nMonthly", f"MSCI EAFE\nMonthly",
     "Markowitz\nCumulative", "Risk Parity\nCumulative", f"MSCI EAFE\nCumulative"],
    bg=C_TITLE)
ws3.row_dimensions[2].height = 30

# Align all three series on common dates
all_dates = sorted(set(mkw_full.index) | set(rp_full.index) | set(msci_mkw.index))
mkw_cum  = (1 + mkw_full).cumprod()
rp_cum   = (1 + rp_full).cumprod()
msci_cum = (1 + msci_mkw).cumprod()

for i, dt in enumerate(all_dates):
    r    = i + 3
    bg_r = C_ALT if i % 2 == 0 else "FFFFFF"
    m_r  = mkw_full.get(dt,  float("nan"))
    r_r  = rp_full.get(dt,   float("nan"))
    ms_r = msci_mkw.get(dt,  float("nan"))
    m_c  = mkw_cum.get(dt,   float("nan"))
    r_c  = rp_cum.get(dt,    float("nan"))
    ms_c = msci_cum.get(dt,  float("nan"))

    vals = [dt.strftime("%b %Y"),
            f"{m_r:+.2%}"  if not np.isnan(m_r)  else "-",
            f"{r_r:+.2%}"  if not np.isnan(r_r)   else "-",
            f"{ms_r:+.2%}" if not np.isnan(ms_r)  else "-",
            f"{m_c:.4f}"   if not np.isnan(m_c)   else "-",
            f"{r_c:.4f}"   if not np.isnan(r_c)   else "-",
            f"{ms_c:.4f}"  if not np.isnan(ms_c)  else "-"]
    write_data_row(ws3, r, vals, bg=bg_r,
                   aligns=["left","center","center","center","center","center","center"])

# ══════════════════════════════════════════════════════════════════════════
# SHEET 4 — WEIGHTS SIDE BY SIDE
# ══════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Portfolio Weights")
ws4.sheet_view.showGridLines = False
for col, w in zip(range(1,8), [8, 30, 24, 10, 14, 14, 12]):
    ws4.column_dimensions[get_column_letter(col)].width = w

ws4.merge_cells("A1:G1")
t4 = ws4["A1"]
t4.value     = "Current Portfolio Weights — Side by Side"
t4.font      = Font(bold=True, size=13, color="FFFFFF")
t4.fill      = fill(C_HEAD)
t4.alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[1].height = 30

write_header_row(ws4, 2,
    ["Ticker", "Company Name", "Sector", "Country",
     "Markowitz\nWeight", "Risk Parity\nWeight", "In Both?"],
    bg=C_TITLE)
ws4.row_dimensions[2].height = 30

mkw_w_cur = cur_mkw["weights"] if cur_mkw else pd.Series(dtype=float)
rp_w_cur  = cur_rp["weights"]  if cur_rp  else pd.Series(dtype=float)
all_tickers_sorted = sorted(
    set(mkw_w_cur[mkw_w_cur > 1e-4].index) | set(rp_w_cur[rp_w_cur > 1e-4].index),
    key=lambda t: -(mkw_w_cur.get(t, 0) + rp_w_cur.get(t, 0))
)

for i, t in enumerate(all_tickers_sorted):
    r    = i + 3
    wm   = mkw_w_cur.get(t, 0.0)
    wr   = rp_w_cur.get(t,  0.0)
    both = "YES" if wm > 1e-4 and wr > 1e-4 else ""
    bg_r = C_ALT if i % 2 == 0 else "FFFFFF"

    vals = [t, names.get(t, t), sectors.get(t, ""), countries.get(t, ""),
            f"{wm:.1%}" if wm > 1e-4 else "-",
            f"{wr:.1%}" if wr > 1e-4 else "-",
            both]
    write_data_row(ws4, r, vals, bg=bg_r,
                   aligns=["left","left","left","center","center","center","center"])
    if wm > 1e-4:
        ws4.cell(r, 5).fill = fill(C_MKW)
    if wr > 1e-4:
        ws4.cell(r, 6).fill = fill(C_RP)
    if both:
        ws4.cell(r, 7).font = Font(bold=True, color="1F4E79", size=10)

# ══════════════════════════════════════════════════════════════════════════
# SHEET 5 — SECTOR & COUNTRY ALLOCATION
# ══════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Sector & Country")
ws5.sheet_view.showGridLines = False
for col, w in zip(range(1,6), [28, 16, 16, 8, 28]):
    ws5.column_dimensions[get_column_letter(col)].width = w

def sector_alloc(weights):
    d = {}
    for t, ww in weights.items():
        if ww > 1e-4:
            s = sectors.get(t, "Other")
            d[s] = d.get(s, 0) + ww
    return d

def country_alloc(weights):
    d = {}
    for t, ww in weights.items():
        if ww > 1e-4:
            c = countries.get(t, "Other")
            d[c] = d.get(c, 0) + ww
    return d

sec_mkw = sector_alloc(mkw_w_cur)
sec_rp  = sector_alloc(rp_w_cur)
cty_mkw = country_alloc(mkw_w_cur)
cty_rp  = country_alloc(rp_w_cur)
all_secs = sorted(set(sec_mkw) | set(sec_rp))
all_ctys = sorted(set(cty_mkw) | set(cty_rp))

# Sector table
ws5.merge_cells("A1:C1")
s1 = ws5["A1"]
s1.value     = "Sector Allocation"
s1.font      = Font(bold=True, size=13, color="FFFFFF")
s1.fill      = fill(C_HEAD)
s1.alignment = Alignment(horizontal="center", vertical="center")
ws5.row_dimensions[1].height = 28

write_header_row(ws5, 2, ["Sector", "Markowitz", "Risk Parity"], bg=C_TITLE)
for i, sec in enumerate(sorted(all_secs, key=lambda s: -sec_mkw.get(s, 0))):
    r    = i + 3
    bg_r = C_ALT if i % 2 == 0 else "FFFFFF"
    wm   = sec_mkw.get(sec, 0.0)
    wr   = sec_rp.get(sec,  0.0)
    write_data_row(ws5, r, [sec, f"{wm:.1%}", f"{wr:.1%}"],
                   bg=bg_r, aligns=["left","center","center"])
    ws5.cell(r, 2).fill = fill(C_MKW)
    ws5.cell(r, 3).fill = fill(C_RP)

# Country table (column E onwards)
r_off = 1
ws5.merge_cells(f"E{r_off}:G{r_off}")
s2 = ws5[f"E{r_off}"]
s2.value     = "Country Allocation"
s2.font      = Font(bold=True, size=13, color="FFFFFF")
s2.fill      = fill(C_HEAD)
s2.alignment = Alignment(horizontal="center", vertical="center")
ws5.row_dimensions[r_off].height = 28
ws5.column_dimensions["E"].width = 28
ws5.column_dimensions["F"].width = 16
ws5.column_dimensions["G"].width = 16

for col, hdr in zip([5, 6, 7], ["Country", "Markowitz", "Risk Parity"]):
    c = ws5.cell(row=r_off+1, column=col, value=hdr)
    c.font      = Font(bold=True, color="FFFFFF", size=11)
    c.fill      = fill(C_TITLE)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border    = thin_border()

for i, cty in enumerate(sorted(all_ctys, key=lambda c: -cty_mkw.get(c, 0))):
    r    = r_off + 2 + i
    bg_r = C_ALT if i % 2 == 0 else "FFFFFF"
    wm   = cty_mkw.get(cty, 0.0)
    wr   = cty_rp.get(cty,  0.0)
    for col, val, al in zip([5, 6, 7],
                             [cty, f"{wm:.1%}", f"{wr:.1%}"],
                             ["left", "center", "center"]):
        c = ws5.cell(row=r, column=col, value=val)
        c.font      = Font(size=10)
        c.alignment = Alignment(horizontal=al, vertical="center")
        c.border    = thin_border()
        c.fill      = fill(bg_r)
    ws5.cell(r, 6).fill = fill(C_MKW)
    ws5.cell(r, 7).fill = fill(C_RP)

wb.save("Z_comparison_report.xlsx")
print("  comparison_report.xlsx saved.")

print(f"\nAll done. Output files:")
print("  portfolio_results.xlsx")
print("  comparison_report.xlsx")
print("  backtest_results.csv")
print("  chart_markowitz_vs_msci.png")
print("  chart_riskparity_vs_msci.png")
