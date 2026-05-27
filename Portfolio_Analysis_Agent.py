# -*- coding: utf-8 -*-
"""
Z_comparison_agent.py
─────────────────────────────────────────────────────────────────────────────
Reads outputs from Z_portfolio_construction_agent.py and the LSEG universe file.
Computes:
  - Financial metrics vs MSCI EAFE benchmark (returns, vol, Sharpe, drawdown,
    beta, calendar-year, 1/3/5yr annualised)
  - ESG weighted-average metrics for each portfolio vs universe equal-weight
  - Top-10 holdings per portfolio (with BICS level 3 sector)
  - Sector allocation at BICS level 3

Saves everything to Z_comparison_results.xlsx for make_factsheets.py.
"""

import sys, io, warnings
if hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
warnings.filterwarnings("ignore")

import numpy as np
import pandas as pd
import openpyxl

RF = 0.025   # risk-free rate

# ─────────────────────────────────────────────────────────────────────────────
# STEP 1 — LOAD INPUTS
# ─────────────────────────────────────────────────────────────────────────────
print("=" * 60)
print("  STEP 1 — Loading inputs")
print("=" * 60)

# Portfolio results: weights + all company data + LSEG metrics (40 companies)
print("  Loading Z_portfolio_results.xlsx...")
df_port = pd.read_excel("Z_portfolio_results.xlsx", engine="openpyxl")
print(f"  {len(df_port)} companies, {len(df_port.columns)} columns")

# Monthly backtest return series produced by Z_portfolio_construction_agent.py
print("  Loading Z_monthly_returns.xlsx...")
df_mon = pd.read_excel("Z_monthly_returns.xlsx", engine="openpyxl",
                       index_col=0, parse_dates=True)
df_mon.index = df_mon.index.to_period("M").to_timestamp()
mkw_full  = df_mon["Markowitz"].dropna()
rp_full   = df_mon["RiskParity"].dropna()
msci_full = df_mon["MSCI_EAFE"].dropna()
print(f"  {len(df_mon)} months  "
      f"({df_mon.index[0].strftime('%b %Y')} → {df_mon.index[-1].strftime('%b %Y')})")

# LSEG universe (~14 000 companies)
print("  Loading LSEG universe...")
wb_l  = openpyxl.load_workbook("Z_master_transition_project_dataset (2).xlsx",
                                read_only=True)
ws_l  = wb_l.active
lhdrs = [ws_l.cell(1, c).value for c in range(1, ws_l.max_column + 1)]
lrows = [list(r) for r in ws_l.iter_rows(min_row=2, values_only=True)]
wb_l.close()
lseg_uni = pd.DataFrame(lrows, columns=lhdrs)
print(f"  {len(lseg_uni)} companies in universe")

# ─────────────────────────────────────────────────────────────────────────────
# STEP 2 — FINANCIAL METRICS vs BENCHMARK
# ─────────────────────────────────────────────────────────────────────────────
print("\n" + "=" * 60)
print("  STEP 2 — Financial metrics vs MSCI EAFE")
print("=" * 60)

def compute_metrics(s, rf=RF):
    n = len(s)
    if n == 0:
        return {k: float("nan") for k in ["cum","ann_r","ann_v","sharpe","max_dd"]}
    cum    = (1 + s).prod() - 1
    ann_r  = (1 + cum) ** (12 / n) - 1
    ann_v  = s.std() * np.sqrt(12)
    sharpe = (ann_r - rf) / ann_v if ann_v > 0 else 0.0
    curve  = (1 + s).cumprod()
    max_dd = ((curve - curve.cummax()) / curve.cummax()).min()
    return dict(cum=cum, ann_r=ann_r, ann_v=ann_v, sharpe=sharpe, max_dd=max_dd)

def compute_beta(p, b):
    df = pd.concat([p, b], axis=1).dropna()
    if len(df) < 2:
        return float("nan")
    p_, b_ = df.iloc[:, 0].values, df.iloc[:, 1].values
    return np.cov(p_, b_)[0, 1] / np.var(b_) if np.var(b_) > 0 else float("nan")

def cal_year(s, yr):
    sub = s[s.index.year == yr]
    return (1 + sub).prod() - 1 if len(sub) > 0 else float("nan")

def ann_window(s, months):
    if len(s) < months:
        return float("nan")
    sub = s.iloc[-months:]
    cum = (1 + sub).prod() - 1
    return (1 + cum) ** (12 / months) - 1

# Align benchmark to each portfolio's test dates
mkw_msci = msci_full.reindex(mkw_full.index).fillna(0)
rp_msci  = msci_full.reindex(rp_full.index).fillna(0)

met_mkw  = compute_metrics(mkw_full)
met_rp   = compute_metrics(rp_full)
met_msci = compute_metrics(msci_full)

beta_mkw = compute_beta(mkw_full, mkw_msci)
beta_rp  = compute_beta(rp_full,  rp_msci)

years = [2021, 2022, 2023, 2024, 2025]
cal_mkw  = {y: cal_year(mkw_full,  y) for y in years}
cal_rp   = {y: cal_year(rp_full,   y) for y in years}
cal_msci = {y: cal_year(msci_full, y) for y in years}

ann1_mkw  = ann_window(mkw_full,  12)
ann3_mkw  = ann_window(mkw_full,  36)
ann5_mkw  = met_mkw["ann_r"]
ann1_rp   = ann_window(rp_full,   12)
ann3_rp   = ann_window(rp_full,   36)
ann5_rp   = met_rp["ann_r"]
ann1_msci = ann_window(msci_full, 12)
ann3_msci = ann_window(msci_full, 36)
ann5_msci = met_msci["ann_r"]

print(f"  Markowitz  : Ann Ret={met_mkw['ann_r']:.1%}  Vol={met_mkw['ann_v']:.1%}"
      f"  Sharpe={met_mkw['sharpe']:.2f}  Beta={beta_mkw:.2f}")
print(f"  Risk Parity: Ann Ret={met_rp['ann_r']:.1%}  Vol={met_rp['ann_v']:.1%}"
      f"  Sharpe={met_rp['sharpe']:.2f}  Beta={beta_rp:.2f}")
print(f"  MSCI EAFE  : Ann Ret={met_msci['ann_r']:.1%}  Vol={met_msci['ann_v']:.1%}"
      f"  Sharpe={met_msci['sharpe']:.2f}")

# ─────────────────────────────────────────────────────────────────────────────
# STEP 3 — ESG METRICS vs UNIVERSE
# ─────────────────────────────────────────────────────────────────────────────
print("\n" + "=" * 60)
print("  STEP 3 — ESG metrics vs universe")
print("=" * 60)

LSEG_COLS = [c for c in df_port.columns if str(c).startswith("LSEG")]

# Universe equal-weighted averages across all ~14 000 companies
uni_avgs = {}
for col in LSEG_COLS:
    raw = pd.to_numeric(lseg_uni[col].replace("NULL", None), errors="coerce")
    if col == "LSEG Biodiversity Due Diligence":
        uni_avgs[col] = (raw > 0).mean() * 100      # % of companies
    else:
        uni_avgs[col] = raw.dropna().mean()

# Portfolio weighted averages (weights already in Z_portfolio_results.xlsx)
def port_wavg(weights_col, cagr_exclude=None):
    result = {}
    for col in LSEG_COLS:
        vals = pd.to_numeric(df_port[col], errors="coerce")
        w    = df_port[weights_col].copy()
        if col == "LSEG Biodiversity Due Diligence":
            has_bio       = (vals > 0)
            active        = w > 1e-4
            result[col]   = has_bio[active].mean() * 100 if active.any() else float("nan")
        else:
            if cagr_exclude and "CAGR" in col:
                mask = ~df_port["ticker"].isin(cagr_exclude)
                vals, w = vals[mask], w[mask]
            valid = vals.notna()
            denom = w[valid].sum()
            result[col] = (w[valid] * vals[valid]).sum() / denom if denom > 0 else float("nan")
    return result

mkw_esg = port_wavg("Weight_Markowitz",  cagr_exclude=["8015"])
rp_esg  = port_wavg("Weight_RiskParity", cagr_exclude=["8015", "AIA"])

print(f"  {'Metric':<55} {'MKW':>7}  {'RP':>7}  {'Universe':>8}")
print(f"  {'-'*80}")
for col in LSEG_COLS:
    print(f"  {col[:53]:<55} {mkw_esg[col]:>7.2f}  {rp_esg[col]:>7.2f}  {uni_avgs[col]:>8.2f}")

# ─────────────────────────────────────────────────────────────────────────────
# STEP 4 — HOLDINGS & SECTOR ALLOCATION (BICS level 3)
# ─────────────────────────────────────────────────────────────────────────────
print("\n" + "=" * 60)
print("  STEP 4 — Holdings & BICS level 3 sector allocation")
print("=" * 60)

def get_holdings(weights_col):
    cols = ["ticker", "company_name", "bics_level_1_name", "bics_level_3_name",
            "country_of_incorporation", weights_col]
    df = df_port[cols].copy()
    df = df[df[weights_col] > 1e-4].sort_values(weights_col, ascending=False).reset_index(drop=True)
    df.index = df.index + 1
    df.index.name = "Rank"
    return df

def sector_bics3(weights_col):
    df = df_port[["bics_level_3_name", weights_col]].copy()
    df = df[df[weights_col] > 1e-4]
    return df.groupby("bics_level_3_name")[weights_col].sum().sort_values(ascending=False)

mkw_holdings = get_holdings("Weight_Markowitz")
rp_holdings  = get_holdings("Weight_RiskParity")
mkw_sec3     = sector_bics3("Weight_Markowitz")
rp_sec3      = sector_bics3("Weight_RiskParity")

print(f"  Markowitz  : {len(mkw_holdings)} holdings, {len(mkw_sec3)} BICS-3 sectors")
print(f"  Risk Parity: {len(rp_holdings)} holdings, {len(rp_sec3)} BICS-3 sectors")

# ─────────────────────────────────────────────────────────────────────────────
# STEP 5 — SAVE Z_comparison_results.xlsx
# ─────────────────────────────────────────────────────────────────────────────
print("\n" + "=" * 60)
print("  STEP 5 — Saving Z_comparison_results.xlsx")
print("=" * 60)

with pd.ExcelWriter("Z_comparison_results.xlsx", engine="openpyxl") as writer:

    # ── Sheet 1: Financial Metrics ────────────────────────────────────────
    fin_df = pd.DataFrame({
        "Metric": [
            "Total Return (5yr)", "Annualised Return", "Annualised Volatility",
            "Sharpe Ratio (rf=2.5%)", "Max Drawdown", "Beta vs MSCI EAFE",
            "1yr Annualised Return", "3yr Annualised Return", "5yr Annualised Return",
        ],
        "Markowitz": [
            met_mkw["cum"], met_mkw["ann_r"], met_mkw["ann_v"],
            met_mkw["sharpe"], met_mkw["max_dd"], beta_mkw,
            ann1_mkw, ann3_mkw, ann5_mkw,
        ],
        "Risk_Parity": [
            met_rp["cum"], met_rp["ann_r"], met_rp["ann_v"],
            met_rp["sharpe"], met_rp["max_dd"], beta_rp,
            ann1_rp, ann3_rp, ann5_rp,
        ],
        "MSCI_EAFE": [
            met_msci["cum"], met_msci["ann_r"], met_msci["ann_v"],
            met_msci["sharpe"], met_msci["max_dd"], 1.0,
            ann1_msci, ann3_msci, ann5_msci,
        ],
    })
    fin_df.to_excel(writer, sheet_name="Financial_Metrics", index=False)

    # ── Sheet 2: Calendar Year Returns ───────────────────────────────────
    cal_df = pd.DataFrame({
        "Year":       years,
        "Markowitz":  [cal_mkw[y]  for y in years],
        "Risk_Parity":[cal_rp[y]   for y in years],
        "MSCI_EAFE":  [cal_msci[y] for y in years],
    })
    cal_df.to_excel(writer, sheet_name="Calendar_Returns", index=False)

    # ── Sheet 3: Monthly Returns (for chart building) ─────────────────────
    mon_df = pd.DataFrame({
        "Markowitz":  mkw_full,
        "RiskParity": rp_full,
        "MSCI_EAFE":  msci_full,
    }).sort_index()
    mon_df.index.name = "Date"
    mon_df.to_excel(writer, sheet_name="Monthly_Returns")

    # ── Sheet 4: ESG Metrics ──────────────────────────────────────────────
    esg_df = pd.DataFrame({
        "Metric":              LSEG_COLS,
        "Markowitz_Portfolio": [mkw_esg[c]   for c in LSEG_COLS],
        "RiskParity_Portfolio":[rp_esg[c]    for c in LSEG_COLS],
        "Universe_Avg":        [uni_avgs[c]  for c in LSEG_COLS],
    })
    esg_df.to_excel(writer, sheet_name="ESG_Metrics", index=False)

    # ── Sheet 5: Markowitz Holdings ───────────────────────────────────────
    mkw_holdings.to_excel(writer, sheet_name="MKW_Holdings")

    # ── Sheet 6: Risk Parity Holdings ────────────────────────────────────
    rp_holdings.to_excel(writer, sheet_name="RP_Holdings")

    # ── Sheet 7: Sector Allocation BICS Level 3 ───────────────────────────
    all_sec3 = sorted(set(mkw_sec3.index) | set(rp_sec3.index))
    sec_df = pd.DataFrame({
        "BICS_Level_3": all_sec3,
        "Markowitz":    [mkw_sec3.get(s, 0.0) for s in all_sec3],
        "Risk_Parity":  [rp_sec3.get(s,  0.0) for s in all_sec3],
    }).sort_values("Markowitz", ascending=False)
    sec_df.to_excel(writer, sheet_name="Sector_BICS3", index=False)

    # ── Sheet 8: Metadata ─────────────────────────────────────────────────
    meta_df = pd.DataFrame({
        "Key": [
            "Backtest_Period", "Benchmark",
            "MKW_CAGR_Footnote", "RP_CAGR_Footnote",
            "MKW_N_Holdings", "RP_N_Holdings",
            "MKW_N_Sectors_BICS3", "RP_N_Sectors_BICS3",
            "MKW_Max_Weight", "RP_Max_Weight",
        ],
        "Value": [
            "Feb 2021 – Apr 2026", "MSCI EAFE (EFA)",
            "Toyota (8015) excluded from 3Y CAGR: LSEG Scope 3 base-year reclassification created a +341% artefact.",
            "Toyota (8015) and Auckland Airport (AIA) excluded from 3Y CAGR: LSEG Scope 3 base-year reclassification created artefacts of +341% and +145% respectively.",
            len(mkw_holdings), len(rp_holdings),
            len(mkw_sec3), len(rp_sec3),
            f"{mkw_holdings['Weight_Markowitz'].max():.1%}",
            f"{rp_holdings['Weight_RiskParity'].max():.1%}",
        ],
    })
    meta_df.to_excel(writer, sheet_name="Metadata", index=False)

print("  Z_comparison_results.xlsx saved.\n")
print("Sheets written:")
print("  Financial_Metrics  — total return, ann return, vol, Sharpe, drawdown, beta, 1/3/5yr")
print("  Calendar_Returns   — annual performance 2021-2025")
print("  Monthly_Returns    — monthly series for chart building")
print("  ESG_Metrics        — portfolio weighted avg vs universe equal-weight")
print("  MKW_Holdings       — all Markowitz holdings ranked by weight")
print("  RP_Holdings        — all Risk Parity holdings ranked by weight")
print("  Sector_BICS3       — sector allocation at BICS level 3")
print("  Metadata           — benchmark, periods, footnotes, portfolio stats")
print("\nDone.")
